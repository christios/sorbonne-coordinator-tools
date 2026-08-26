"""Read a group-assignment workbook's Reference sheet into scopes, groups and CRNs.

The Reference sheet is the part of those workbooks that everything else derives from:
one row per CRN, saying which group of which block it belongs to. Reading it gives the
Student Database its starting catalogue, after which coordinators maintain it here and
the workbook is no longer needed.

Two shapes exist and both are read:

*Cohort* workbooks (FYS, L1, L2) — a block of courses taught in parallel groups::

    CRN | Group | Scope | Course Code | Course Name | Component | Teacher | … | Helper key
    23223 |   1   |  TD   |  MATH001    | Pre-calculus |    TD    |   Dr …  | … | TD|1|MATH001

  Scope is the block, so group 1 of TD means one bundle: MATH001, MATH009 and MATH011.

*Language* workbooks — one class per group, with a seat limit::

    CRN | Language group | Level | Group | Open to | Day | Time | Teacher | Capacity | …
    23302 |     A0-F1     |  A0   |  F1   |   FYS   | Tue | 4:30 |   Dr …  |    30    | …

  The level is the block and the language group is the group, which holds a single CRN.

Neither shape is guessed at from the filename: the header row names the columns, and a
sheet that matches neither is refused rather than half-read.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO

import openpyxl

# A single-class block still needs a column in the matrix; this is what it is called.
SINGLE_COURSE_CODE = "CLASS"


class ReferenceImportError(Exception):
    """The workbook could not be read, with a sentence a coordinator can act on."""


@dataclass
class ImportedCourse:
    code: str
    name: str = ""
    component: str = ""


@dataclass
class ImportedGroup:
    label: str
    capacity: int = 0
    note: str = ""
    #: course code -> (crn, teacher)
    crns: dict[str, tuple[str, str]] = field(default_factory=dict)


@dataclass
class ImportedScope:
    code: str
    name: str = ""
    #: Which student tab this block's column lives on. Two blocks share a tab when the
    #: coordinator put them there — Readiness sits beside the tutorials, not on its own —
    #: and that is layout the Reference sheet records and nothing else knows.
    tab: str = ""
    #: What that column is called on the tab: "TD group", "Readiness group".
    group_column: str = ""
    #: Which column that is, so two blocks on one tab come back in the order they were in.
    column_index: int = 0
    courses: list[ImportedCourse] = field(default_factory=list)
    groups: list[ImportedGroup] = field(default_factory=list)


@dataclass
class ReferenceImport:
    """What one Reference sheet holds, ready to be stored."""

    scopes: list[ImportedScope] = field(default_factory=list)
    sheet: str = ""
    style: str = ""
    crn_count: int = 0

    @property
    def group_count(self) -> int:
        return sum(len(scope.groups) for scope in self.scopes)


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).split())


def _header_index(sheet) -> tuple[int, dict[str, int]]:
    """Find the row that starts with CRN, and where each of its columns sits."""
    for number, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        labels = [_text(cell).lower() for cell in row]
        if labels and labels[0] == "crn":
            return number, {label: position for position, label in enumerate(labels) if label}
    raise ReferenceImportError(
        f"The sheet {sheet.title!r} has no header row starting with CRN, so it is not a Reference sheet."
    )


def _reference_sheet(book: openpyxl.Workbook):
    for sheet in book.worksheets:
        if sheet.title.strip().lower().startswith("reference"):
            return sheet
    raise ReferenceImportError(
        "That workbook has no Reference sheet. Upload the group-assignment workbook itself, "
        "not a copy of one of its student tabs."
    )


def _scope_of(
    scopes: dict[str, ImportedScope], code: str, name: str, tab: str = "", group_column: str = ""
) -> ImportedScope:
    scope = scopes.get(code)
    if scope is None:
        scope = ImportedScope(code=code, name=name, tab=tab, group_column=group_column)
        scopes[code] = scope
        return scope
    if not scope.name:
        scope.name = name
    # First row wins: the sheet repeats the same tab on every row of a block.
    if not scope.tab:
        scope.tab = tab
    if not scope.group_column:
        scope.group_column = group_column
    return scope


def _course_of(scope: ImportedScope, course: ImportedCourse) -> None:
    if not any(existing.code == course.code for existing in scope.courses):
        scope.courses.append(course)


def _group_of(scope: ImportedScope, label: str) -> ImportedGroup:
    for group in scope.groups:
        if group.label == label:
            return group
    group = ImportedGroup(label=label)
    scope.groups.append(group)
    return group


def _read_cohort_rows(rows, columns) -> dict[str, ImportedScope]:
    scopes: dict[str, ImportedScope] = {}
    for row in rows:
        crn = _text(row[columns["crn"]])
        scope_code = _text(row[columns["scope"]])
        label = _text(row[columns["group"]])
        course_code = _text(row[columns["course code"]])
        if not (crn and scope_code and label and course_code):
            continue
        scope = _scope_of(
            scopes,
            scope_code,
            "",
            tab=_pick(row, columns, "tab"),
            group_column=_pick(row, columns, "group column"),
        )
        _course_of(
            scope,
            ImportedCourse(
                code=course_code,
                name=_pick(row, columns, "course name"),
                component=_pick(row, columns, "component"),
            ),
        )
        _group_of(scope, label).crns[course_code] = (crn, _pick(row, columns, "teacher"))
    return scopes


def _read_language_rows(rows, columns) -> dict[str, ImportedScope]:
    scopes: dict[str, ImportedScope] = {}
    for row in rows:
        crn = _text(row[columns["crn"]])
        level = _text(row[columns["level"]])
        label = _pick(row, columns, "language group") or _pick(row, columns, "group")
        if not (crn and level and label):
            continue
        scope = _scope_of(scopes, level, f"Level {level}")
        # One class per group, so the matrix is a single column.
        _course_of(scope, ImportedCourse(code=SINGLE_COURSE_CODE, name=scope.name))
        group = _group_of(scope, label)
        group.crns[SINGLE_COURSE_CODE] = (crn, _pick(row, columns, "teacher"))
        group.capacity = _number(_pick(row, columns, "capacity"))
        group.note = " · ".join(
            part
            for part in (
                _pick(row, columns, "day"),
                _pick(row, columns, "time"),
                _open_to(_pick(row, columns, "open to")),
            )
            if part
        )
    return scopes


def _open_to(value: str) -> str:
    return f"open to {value}" if value else ""


def _pick(row, columns: dict[str, int], label: str) -> str:
    position = columns.get(label)
    if position is None or position >= len(row):
        return ""
    return _text(row[position])


def _number(value: str) -> int:
    try:
        return max(0, int(float(value)))
    except (TypeError, ValueError):
        return 0


def parse_group_reference(content: bytes, filename: str = "") -> ReferenceImport:
    """Read one workbook's Reference sheet. Raises ReferenceImportError with a reason."""
    try:
        book = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise ReferenceImportError(
            f"{filename or 'That file'} could not be read as an Excel workbook."
        ) from exc

    sheet = _reference_sheet(book)
    header_row, columns = _header_index(sheet)
    rows = list(sheet.iter_rows(min_row=header_row + 1, values_only=True))

    if "scope" in columns and "course code" in columns:
        style, scopes = "cohort", _read_cohort_rows(rows, columns)
    elif "level" in columns:
        style, scopes = "language", _read_language_rows(rows, columns)
    else:
        raise ReferenceImportError(
            f"The Reference sheet in {filename or 'that workbook'} has neither a Scope column nor a "
            "Level column, so its groups cannot be read."
        )

    if not scopes:
        raise ReferenceImportError(
            f"The Reference sheet in {filename or 'that workbook'} has a header but no CRNs under it."
        )

    ordered = sorted(scopes.values(), key=lambda scope: scope.code)
    for scope in ordered:
        scope.groups.sort(key=lambda group: _sortable(group.label))
        scope.courses.sort(key=lambda course: course.code)

    return ReferenceImport(
        scopes=ordered,
        sheet=sheet.title,
        style=style,
        crn_count=sum(len(group.crns) for scope in ordered for group in scope.groups),
    )


def _sortable(label: str) -> tuple:
    """Order group labels the way a person reads them: 1, 2, 10 — and 1A before 1B."""
    digits, rest = "", ""
    for character in label:
        if character.isdigit() and not rest:
            digits += character
        else:
            rest += character
    return (int(digits) if digits else 0, rest.lower(), label)
