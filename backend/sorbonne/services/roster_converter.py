from __future__ import annotations

from dataclasses import asdict, dataclass
from io import BytesIO
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


class ConversionError(Exception):
    """Raised when a PDF cannot be converted as a Course Class Roster."""


@dataclass(frozen=True)
class CourseInfo:
    term: str | None = None
    crn: str | None = None
    course_code: str | None = None
    course_title: str | None = None
    department: str | None = None
    contact_hours: str | None = None
    teacher: str | None = None
    major: str | None = None
    printed_on: str | None = None
    source_filename: str | None = None


@dataclass
class RosterRow:
    number: int
    student_id: str
    student_name: str
    status: str
    major: str
    department: str
    level: str
    hours_of_absences: float
    absence_percent: str


@dataclass
class RosterConversionResult:
    course_info: CourseInfo
    rows: list[RosterRow]
    page_count: int

    @property
    def excel_filename(self) -> str:
        code = self.course_info.course_code or "course-roster"
        crn = self.course_info.crn
        parts = [code]
        if crn:
            parts.append(crn)
        parts.append("roster.xlsx")
        return _safe_filename("-".join(parts))

    def to_preview_payload(self) -> dict[str, Any]:
        return {
            "courseInfo": asdict(self.course_info),
            "rows": [asdict(row) for row in self.rows],
            "pageCount": self.page_count,
            "rowCount": len(self.rows),
            "excelFilename": self.excel_filename,
        }

    def to_xlsx_bytes(self) -> bytes:
        workbook = Workbook()
        info_sheet = workbook.active
        info_sheet.title = "Course Info"
        roster_sheet = workbook.create_sheet("Roster")

        _write_course_info_sheet(info_sheet, self.course_info, len(self.rows), self.page_count)
        _write_roster_sheet(roster_sheet, self.rows)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()


_ROW_RE = re.compile(
    r"^\s*(?P<number>\d+)\s+"
    r"(?P<student_id>A\d{8})\s+"
    r"(?P<student_name>.+?)\s+"
    r"(?P<status>[A-Z]{1,4})\s+"
    r"(?P<major>\S+)\s+"
    r"(?P<department>\S+)\s+"
    r"(?P<level>\S+)\s+"
    r"(?P<hours>\d+(?:\.\d+)?)\s+"
    r"(?P<absence_percent>\d+(?:\.\d+)?%)\s*$"
)

_SKIP_LINE_PREFIXES = (
    "#",
    "Course Class Roster",
    "Second Semester",
    "Page ",
    "Printed On",
)


def convert_course_roster_pdf(pdf_content: bytes, source_filename: str | None = None) -> RosterConversionResult:
    try:
        reader = PdfReader(BytesIO(pdf_content))
    except Exception as exc:
        raise ConversionError("This file could not be opened as a PDF.") from exc

    if reader.is_encrypted:
        raise ConversionError("Encrypted PDFs are not supported yet.")

    page_texts = [_extract_page_text(page) for page in reader.pages]
    full_text = "\n".join(page_texts)
    rows = _extract_rows(page_texts)
    if not rows:
        raise ConversionError("No Course Class Roster rows were found in this PDF.")

    missing_numbers = _missing_sequence_numbers(rows)
    if missing_numbers:
        missing = ", ".join(str(number) for number in missing_numbers[:10])
        raise ConversionError(f"The roster table appears incomplete. Missing row numbers: {missing}.")

    return RosterConversionResult(
        course_info=_extract_course_info(full_text, source_filename),
        rows=rows,
        page_count=len(reader.pages),
    )


def _extract_page_text(page: Any) -> str:
    try:
        return page.extract_text(extraction_mode="layout") or ""
    except TypeError:
        return page.extract_text() or ""


def _extract_rows(page_texts: list[str]) -> list[RosterRow]:
    rows: list[RosterRow] = []
    current: RosterRow | None = None

    for text in page_texts:
        for line in text.splitlines():
            match = _ROW_RE.match(line)
            if match:
                data = match.groupdict()
                current = RosterRow(
                    number=int(data["number"]),
                    student_id=data["student_id"],
                    student_name=_normalize_spaces(data["student_name"]),
                    status=data["status"],
                    major=data["major"],
                    department=data["department"],
                    level=data["level"],
                    hours_of_absences=_parse_number(data["hours"]),
                    absence_percent=data["absence_percent"],
                )
                rows.append(current)
                continue

            if current is None:
                continue

            continuation = line.strip()
            if _is_name_continuation(continuation):
                current.student_name = _normalize_spaces(f"{current.student_name} {continuation}")

    return rows


def _extract_course_info(text: str, source_filename: str | None) -> CourseInfo:
    metadata_text = re.split(r"\n\s*#\s+ID\s+Student Name", text, maxsplit=1)[0]
    return CourseInfo(
        term=_field(metadata_text, "Term", "CRN"),
        crn=_field(metadata_text, "CRN"),
        course_code=_field(metadata_text, "Course Code", "Course Title"),
        course_title=_field(metadata_text, "Course Title"),
        department=_field(metadata_text, "Department", "Contact Hours"),
        contact_hours=_field(metadata_text, "Contact Hours"),
        teacher=_field(metadata_text, "Teacher", "Major"),
        major=_field(metadata_text, "Major"),
        printed_on=_printed_on(text),
        source_filename=source_filename,
    )


def _field(text: str, label: str, next_label: str | None = None) -> str | None:
    if next_label:
        pattern = rf"{re.escape(label)}[ \t]+(?P<value>.+?)[ \t]+{re.escape(next_label)}\b"
    else:
        pattern = rf"{re.escape(label)}[ \t]+(?P<value>[^\n\r]+)"

    match = re.search(pattern, text)
    if not match:
        return None
    return _normalize_spaces(match.group("value"))


def _printed_on(text: str) -> str | None:
    match = re.search(r"Printed On\s+(?P<printed_on>[^\n\r]+)", text)
    if not match:
        return None
    return _normalize_spaces(match.group("printed_on"))


def _is_name_continuation(line: str) -> bool:
    if not line:
        return False
    if any(line.startswith(prefix) for prefix in _SKIP_LINE_PREFIXES):
        return False
    if "Absences" in line or "Absence %" in line:
        return False
    return bool(re.search(r"[A-Za-z]", line)) and not bool(re.search(r"\b(A\d{8}|RE|MATH|SCEN|UG)\b", line))


def _missing_sequence_numbers(rows: list[RosterRow]) -> list[int]:
    numbers = {row.number for row in rows}
    first = min(numbers)
    last = max(numbers)
    return [number for number in range(first, last + 1) if number not in numbers]


def _normalize_spaces(value: str | None) -> str | None:
    if value is None:
        return None
    return re.sub(r"\s+", " ", value).strip()


def _parse_number(value: str) -> int | float:
    numeric = float(value)
    if numeric.is_integer():
        return int(numeric)
    return numeric


def _safe_filename(value: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9_.-]+", "-", value)
    return re.sub(r"-+", "-", safe).strip("-")


def _write_course_info_sheet(sheet: Any, course_info: CourseInfo, row_count: int, page_count: int) -> None:
    sheet.append(["Field", "Value"])
    fields = [
        ("Term", course_info.term),
        ("CRN", course_info.crn),
        ("Course Code", course_info.course_code),
        ("Course Title", course_info.course_title),
        ("Department", course_info.department),
        ("Contact Hours", course_info.contact_hours),
        ("Teacher", course_info.teacher),
        ("Major", course_info.major),
        ("Printed On", course_info.printed_on),
        ("Source File", course_info.source_filename),
        ("Rows Converted", row_count),
        ("Pages", page_count),
    ]
    for field, value in fields:
        sheet.append([field, value or ""])

    _style_header(sheet, 1, 2)
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 48
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_roster_sheet(sheet: Any, rows: list[RosterRow]) -> None:
    headers = [
        "#",
        "ID",
        "Student Name",
        "Status",
        "Major",
        "Dept.",
        "Level",
        "#Hours of Absences",
        "Absence %",
    ]
    sheet.append(headers)

    for row in rows:
        sheet.append(
            [
                row.number,
                row.student_id,
                row.student_name,
                row.status,
                row.major,
                row.department,
                row.level,
                row.hours_of_absences,
                row.absence_percent,
            ]
        )

    _style_header(sheet, 1, len(headers))
    widths = [8, 14, 42, 10, 12, 12, 10, 20, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column == 3)
            cell.border = Border(bottom=Side(style="thin", color="E5E7EB"))


def _style_header(sheet: Any, row_number: int, column_count: int) -> None:
    fill = PatternFill(fill_type="solid", fgColor="D9E2F3")
    font = Font(bold=True, color="111827")
    border = Border(bottom=Side(style="thin", color="94A3B8"))
    for column in range(1, column_count + 1):
        cell = sheet.cell(row=row_number, column=column)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
