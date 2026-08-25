"""Read which group each student was put in, from the workbooks coordinators already fill.

The same group-assignment templates that seed the catalogue (see `group_reference_import`)
also carry the answer to "who is in which group": one row per student, one typed group per
block. The CRN columns beside them are Excel formulas, and every one has the same shape::

    =IF($E2="","",IFERROR(INDEX(FYS_CRN,MATCH("TD|"&$E2&"|MATH001",FYS_KEY,0)),"group?"))

That formula is the map. `"TD|"` names the block, `$E2` says which column the coordinator
types the group into, and the rest is the course. So the workbook tells us where its own
groups live rather than us guessing from header text — a renamed column or a new course tab
goes on working.

Unlike the student platform's reader, this one deliberately stops at the group. CRNs belong
to the catalogue here, which is validated against the timetable; resolving them twice, from
two sources, is exactly the drift this whole change exists to remove.

The derived cells are no use: openpyxl cannot evaluate formulas, and a template that was
never opened in Excel carries no cached values. So the formulas are read, not the results.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO

import openpyxl

STUDENT_ID_HEADERS = ("student id", "banner id", "id")

# MATCH("TD|"&$E2&"|MATH001", FYS_KEY, 0) — the block, the column holding the group.
LOOKUP_PATTERN = re.compile(
    r'MATCH\(\s*"(?P<prefix>[^"]*)"\s*&\s*\$?(?P<column>[A-Z]{1,3})\$?\d+',
    re.IGNORECASE,
)


class AssignmentImportError(Exception):
    """Raised when a workbook cannot be read as a group-assignment template."""


@dataclass
class AssignmentImport:
    """`students` is `{student id: {scope code: group label}}`."""

    students: dict[str, dict[str, str]] = field(default_factory=dict)
    sheets_read: list[str] = field(default_factory=list)
    scopes_seen: set[str] = field(default_factory=set)
    blank_rows: int = 0

    @property
    def assignment_count(self) -> int:
        return sum(len(groups) for groups in self.students.values())


def _text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def normalize_student_id(value: object) -> str:
    """The same normalisation the student lookup uses, so the two agree on who is who."""
    return re.sub(r"[^A-Z0-9]", "", _text(value).upper())


def _student_heading(sheet) -> tuple[int, int] | None:
    """Where the ids are: `(row of the header, its column)`. Students start below it."""
    for row in sheet.iter_rows(min_row=1, max_row=8):
        for cell in row:
            if _text(cell.value).lower() in STUDENT_ID_HEADERS:
                return cell.row, cell.column
    return None


def _group_columns(sheet) -> dict[int, str]:
    """`column holding the group -> scope code`, read out of the CRN formulas."""
    found: dict[int, str] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or not cell.value.startswith("="):
                continue
            match = LOOKUP_PATTERN.search(cell.value)
            if match is None:
                continue
            scope = match.group("prefix").strip().strip("|").strip()
            if not scope:
                continue
            column = openpyxl.utils.column_index_from_string(match.group("column"))
            found.setdefault(column, scope.upper())
    return found


def parse_group_assignments(content: bytes, filename: str = "") -> AssignmentImport:
    """Every student's typed group, per block, across a workbook's sheets."""
    try:
        book = openpyxl.load_workbook(BytesIO(content), data_only=False)
    except Exception as exc:
        raise AssignmentImportError("That file could not be read as an Excel workbook.") from exc

    report = AssignmentImport()

    for sheet in book.worksheets:
        heading = _student_heading(sheet)
        group_columns = _group_columns(sheet)
        if heading is None or not group_columns:
            continue
        report.sheets_read.append(sheet.title)
        _read_sheet(sheet, heading, group_columns, report)

    if not report.students:
        raise AssignmentImportError(
            f"No group assignments were found in {filename or 'that workbook'}. Fill in at least "
            "one student's group before uploading it."
        )
    return report


def _read_sheet(
    sheet, heading: tuple[int, int], group_columns: dict[int, str], report: AssignmentImport
) -> None:
    header_row, student_column = heading
    # Start below the header: its own text normalises to something that looks like an id.
    for row in sheet.iter_rows(min_row=header_row + 1):
        cells = {cell.column: cell.value for cell in row}
        student = normalize_student_id(cells.get(student_column))
        if not student:
            continue

        groups: dict[str, str] = {}
        for column, scope in group_columns.items():
            label = _text(cells.get(column))
            if label:
                groups[scope] = label
                report.scopes_seen.add(scope)

        if not groups:
            # A student on the sheet with no group typed anywhere is not an assignment;
            # they are the coordinator's unfinished work, and readiness will say so.
            report.blank_rows += 1
            continue
        report.students.setdefault(student, {}).update(groups)
